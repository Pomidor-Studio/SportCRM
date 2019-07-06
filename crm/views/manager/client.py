import re
from datetime import date, datetime
import pytz

from itertools import chain
import openpyxl as openpyxl
from django.contrib import messages
from django.contrib.auth.views import SuccessURLAllowedHostsMixin
from django.db import transaction
from django.db.models import F
from django.forms import forms
from django.forms.utils import ErrorList
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils.http import is_safe_url
from django.views.generic import (
    CreateView, DeleteView, FormView, RedirectView, TemplateView,
    UpdateView,
)
from django.contrib.auth import REDIRECT_FIELD_NAME
from django_filters.views import FilterView
from django_multitenant.utils import get_current_tenant
from openpyxl.utils import cell
from rest_framework.generics import RetrieveAPIView
from rest_framework.serializers import DateField, IntegerField
from reversion.views import RevisionMixin
from rules.contrib.views import PermissionRequiredMixin

from crm import utils
from crm.enums import BALANCE_REASON
from crm.filters import ClientFilter
from crm.forms import (
    ClientForm, ClientSubscriptionForm, ExtendClientSubscriptionForm,
    UploadExcelForm,
)
from crm.models import (
    Attendance, Client, ClientSubscriptions, EventClass, ExtensionHistory,
    SubscriptionsType, Event, ExtensionHistory
)
from crm.serializers import ClientSubscriptionCheckOverlappingSerializer
from crm.templatetags.html_helper import (
    allowed_date_formats_ru,
    get_vk_user_ids, try_parse_date,
)
from crm.views.manager.event_class import EventByDateMixin
from crm.views.mixin import CreateAndAddView, UnDeleteView
from gcp.tasks import enqueue


class List(PermissionRequiredMixin, FilterView):
    model = Client
    filterset_class = ClientFilter
    template_name = 'crm/manager/client/list.html'
    context_object_name = 'clients'
    paginate_by = 1000
    permission_required = 'client'
    ordering = ['-deleted', 'name']

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=object_list, **kwargs)
        context['has_active_event_class'] = EventClass.objects.active().exists()
        context['vk_group_id'] = get_current_tenant().vk_group_id
        all_clients = Client.objects.all()
        context['all_clients_count'] = all_clients.count
        debtor_count = 0
        long_time_not_go_count = 0
        for client in all_clients:
            if ClientFilter.filter_debtor(self,
                                          queryset=Client.objects.filter(name=client.name),
                                          name=None,
                                          value=None):
                debtor_count = debtor_count + 1
            if ClientFilter.filter_long_time_not_go(self,
                                                    queryset=Client.objects.filter(name=client.name),
                                                    name=None,
                                                    value=None):
                long_time_not_go_count += long_time_not_go_count + 1
        context['debtor_count'] = debtor_count
        context['long_time_not_go_count'] = long_time_not_go_count
        return context


class ClientEditMixin(SuccessURLAllowedHostsMixin):

    def get_redirect_url(self):
        redirect_to = self.request.POST.get(
            REDIRECT_FIELD_NAME,
            self.request.GET.get(REDIRECT_FIELD_NAME, '')
        )
        url_is_safe = is_safe_url(
            url=redirect_to,
            allowed_hosts=self.get_success_url_allowed_hosts(),
            require_https=self.request.is_secure(),
        )
        return redirect_to if url_is_safe else ''

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        next_page = self.get_redirect_url()
        if next_page:
            context[REDIRECT_FIELD_NAME] = next_page
        return context

    def get_success_url(self):
        next_page = self.get_redirect_url()
        if next_page:
            return next_page
        return super().get_success_url()


class Create(PermissionRequiredMixin, RevisionMixin, ClientEditMixin, CreateAndAddView):
    model = Client
    form_class = ClientForm
    template_name = 'crm/manager/client/form.html'
    permission_required = 'client.add'
    add_another_url = 'crm:manager:client:new'
    message_info = 'Ученик успешно создан'


class UnMarkClient(
    PermissionRequiredMixin,
    RevisionMixin,
    EventByDateMixin,
    RedirectView
):
    permission_required = 'event.mark-attendance'

    def get(self, request, *args, **kwargs):
        event = self.get_object()
        client_id = self.kwargs.pop('client_id')
        self.kwargs.setdefault('pk', client_id)
        client = Client.objects.get(id=client_id)
        client.restore_visit(event)
        self.kwargs.pop('event_class_id')
        self.kwargs.pop('year')
        self.kwargs.pop('month')
        self.kwargs.pop('day')
        self.url = self.get_success_url()
        messages.info(self.request, 'Посещение мероприятия удалено')
        return super().get(request, *args, **kwargs)

    def get_success_url(self):
        return reverse(
            'crm:manager:client:detail', kwargs=self.kwargs)


class CancelAttendance(
    PermissionRequiredMixin,
    RevisionMixin,
    EventByDateMixin,
    RedirectView
):
    permission_required = 'event.manipulate'

    def get(self, request, *args, **kwargs):
        event = self.get_object()
        client_id = self.kwargs.pop('client_id')
        self.kwargs.setdefault('pk', client_id)
        client = Client.objects.get(id=client_id)
        attendance = Attendance.objects.get(client_id=client.id, event_id=event.id)
        client.cancel_signup_for_event(event)
        self.kwargs.pop('event_class_id')
        self.kwargs.pop('year')
        self.kwargs.pop('month')
        self.kwargs.pop('day')
        self.url = self.get_success_url()
        if attendance.signed_up and not attendance.marked:
            messages.info(self.request, 'Запись на мероприятие удалена')
        else:
            messages.info(self.request, 'Посещение мероприятия удалено')
        return super().get(request, *args, **kwargs)

    def get_success_url(self):
        return reverse(
            'crm:manager:client:detail', kwargs=self.kwargs)


class Update(PermissionRequiredMixin, RevisionMixin, ClientEditMixin, UpdateView):
    model = Client
    form_class = ClientForm
    template_name = 'crm/manager/client/form.html'
    permission_required = 'client.edit'

    def get_initial(self):
        initial = super().get_initial()
        vkid = self.object.vk_user_id
        initial['vk_page'] = 'https://vk.com/id{}'.format(vkid) if vkid else None
        return initial


class Delete(PermissionRequiredMixin, RevisionMixin, DeleteView):
    model = Client
    template_name = 'crm/manager/client/confirm_delete.html'
    success_url = reverse_lazy('crm:manager:client:list')
    permission_required = 'client.delete'


class UnDelete(PermissionRequiredMixin, RevisionMixin, UnDeleteView):
    model = Client
    success_url = reverse_lazy('crm:manager:client:list')
    template_name = 'crm/manager/client/confirm_undelete.html'
    permission_required = 'client.undelete'


class AddSubscription(
    PermissionRequiredMixin,
    RevisionMixin,
    CreateView
):
    form_class = ClientSubscriptionForm
    template_name = "crm/manager/client/add-subscriptions.html"
    permission_required = 'client_subscription.sale'

    def get_client(self) -> Client:
        try:
            return Client.all_objects.get(id=self.kwargs['client_id'])
        except Client.DoesNotExist:
            raise Http404('No client matches the given query.')

    def get_event(self) -> Event:
        try:
            event_date = date(
                self.kwargs['year'], self.kwargs['month'], self.kwargs['day']
            )
            return Event.objects.get_or_virtual(
                self.kwargs['event_class_id'], event_date
            )
        except Exception:
            pass

    def get_success_url(self):
        if self.form.cleaned_data.get('go_back'):
            return self.form.cleaned_data['go_back']

        event = self.get_event()
        if event:
            args = (
                event.event_class.id,
                event.date.year,
                event.date.month,
                event.date.day,
            )
            return reverse(
                'crm:manager:event-class:event:event-by-date', args=args
            )

        return reverse(
            'crm:manager:client:detail', args=[self.kwargs['client_id']])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        event = self.get_event()
        if event:
            kwargs['event_class'] = event.event_class
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        client = self.get_client()
        initial['client'] = client
        initial['sold_by'] = self.request.user
        initial['go_back'] = self.request.GET.get('gb')

        # Add preselected subscription type from previous client subscriptions
        # history.
        last_sub = client.last_sub()
        if last_sub:
            initial['subscription'] = last_sub.subscription
            initial['visits_left'] = last_sub.subscription.visit_limit
            initial['price'] = last_sub.subscription.price
        event = self.get_event()
        if event:
            initial['start_date'] = event.date
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        client = self.get_client()
        context['client'] = client
        context['allow_check_overlapping'] = True

        attendance = client.attendance_set.annotate(
            sort_dt=F('event__date')
        ).order_by('-event__date')
        for a in attendance:
            tm = a.event.start_time if a.event.start_time else datetime.min.time()
            a.sort_dt = datetime.combine(a.sort_dt, tm).replace(tzinfo=pytz.UTC)

        extensionhistory = ExtensionHistory.objects.annotate(
            sort_dt=F('date_extended')
        ).filter(
            client_subscription__client__id=client.id,
        ).order_by(
            '-date_extended'
        )

        balancehistory = client.clientbalancechangehistory_set.annotate(
            sort_dt=F('entry_date')
        ).order_by('-entry_date')

        attendance_with_balance = list(chain(attendance, balancehistory, extensionhistory))
        attendance_with_balance = sorted(
            attendance_with_balance,
            key=lambda i: i.sort_dt,
            reverse=True,
        )
        context['attendance_with_balance'] = attendance_with_balance
        context['hide_form'] = self.kwargs.get('hide_form')
        context['event'] = self.get_event()

        return context

    def form_valid(self, form):
        self.form = form
        cash_earned = form.cleaned_data['cash_earned']
        abon_price = form.cleaned_data['price']
        client = self.get_client()
        current_user = self.request.user

        with transaction.atomic():
            client.add_balance_in_history(
                -abon_price, BALANCE_REASON.BY_SUBSCRIPTION,
                skip_notification=True,
                changed_by=current_user,
            )
            if cash_earned:
                client.add_balance_in_history(
                    abon_price, BALANCE_REASON.UPDATE_BALANCE,
                    skip_notification=True,
                    changed_by=current_user,
                )
            client.save()
            response = super().form_valid(form)
            self.object.event = self.get_event()
            self.object.save()
            enqueue('notify_client_buy_subscription', self.object.id)

        return response


class SubscriptionUpdate(
    PermissionRequiredMixin,
    RevisionMixin,
    UpdateView
):
    model = ClientSubscriptions
    form_class = ClientSubscriptionForm
    template_name = 'crm/manager/client/add-subscriptions.html'
    permission_required = 'client_subscription.edit'

    def activated_subscription(self):
        return self.object.attendance_set.exists()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['disable_subscription_type'] = True
        kwargs['activated_subscription'] = self.activated_subscription()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['history'] = ExtensionHistory.objects.filter(
            client_subscription=self.object.id)
        client = self.object.client
        context['client'] = client
        context['activated_subscription'] = self.activated_subscription()
        context['allow_check_overlapping'] = False
        attendance = client.attendance_set.order_by('-event__date')
        balancehistory = client.clientbalancechangehistory_set.order_by(
            '-entry_date')
        # TODO: fix ordering
        attendance_with_balance = chain(attendance, balancehistory)
        context['attendance_with_balance'] = attendance_with_balance
        context['hide_form'] = self.kwargs.get('hide_form')
        if self.kwargs.get('event_class_id'):
            event_class_id = self.kwargs.get('event_class_id')
            context.update(
                event_class_name=EventClass.objects.get(pk=event_class_id).name,
                event_class_id=event_class_id,
                event_year=self.kwargs.get('year'),
                event_month=self.kwargs.get('month'),
                event_day=self.kwargs.get('day'),
            )
        return context

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        if self.activated_subscription():
            messages.warning(
                self.request,
                'По этому абонементу были посещения, по этому, его уже нельзя '
                'редактировать.'
            )
        return response


class AddSubscriptionWithExtending(AddSubscription):
    object: ClientSubscriptions = ...

    def form_valid(self, form):
        with transaction.atomic():
            ret_val = super().form_valid(form)

            to_cancel_events = self.object.canceled_events()[:len(
                self.object.remained_events()
            )]
            for event in to_cancel_events:
                self.object.extend_by_cancellation(event)
        return ret_val


class CheckOverlapping(RetrieveAPIView):
    serializer_class = ClientSubscriptionCheckOverlappingSerializer

    def get_object(self):
        subscription = SubscriptionsType.objects.get(
            id=self.request.query_params.get('st'))
        start_date = DateField(input_formats=['%d.%m.%Y']).to_internal_value(
            self.request.query_params.get('start'))
        visits_left = IntegerField().to_internal_value(
            self.request.query_params.get('vl'))

        return ClientSubscriptions(
            subscription=subscription,
            start_date=start_date,
            end_date=subscription.end_date(start_date),
            visits_left=visits_left
        )


class SubscriptionExtend(PermissionRequiredMixin, RevisionMixin, FormView):
    form_class = ExtendClientSubscriptionForm
    template_name = 'crm/manager/client/subscription_extend.html'
    permission_required = 'client_subscription.extend'

    object: ClientSubscriptions = ...

    def get_initial(self):
        initial = super().get_initial()
        initial['go_back'] = self.request.GET.get('gb')
        return initial

    def get_object(self):
        self.object = get_object_or_404(
            ClientSubscriptions, id=self.kwargs['pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['subscription'] = self.object
        return kwargs

    def get(self, request, *args, **kwargs):
        self.get_object()
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.get_object()
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        self.form = form
        self.object.extend_duration(
            form.cleaned_data['visit_limit'],
            form.cleaned_data['reason']
        )
        return super().form_valid(form)

    def get_success_url(self):
        if self.form.cleaned_data.get('go_back'):
            return self.form.cleaned_data['go_back']

        return reverse(
            'crm:manager:client:detail', args=(self.object.client_id,))


class SubscriptionDelete(PermissionRequiredMixin, RevisionMixin, DeleteView):
    model = ClientSubscriptions
    template_name = 'crm/manager/client/subscription_confirm_delete.html'
    permission_required = 'client_subscription.delete'

    def get_success_url(self):
        return reverse(
            'crm:manager:client:detail', args=[self.object.client.id])


class ImportReport(PermissionRequiredMixin, RevisionMixin, TemplateView):
    template_name = 'crm/manager/client/import_report.html'
    permission_required = 'client.add'

    def get(self, *args, **kwargs):
        if 'import_errors' not in self.request.session:
            raise Http404()
        resp = super().get(*args, **kwargs)
        del self.request.session['import_errors']
        return resp

    def get_context_data(self, **kwargs):
        context = super(ImportReport, self).get_context_data(**kwargs)
        context['import_errors'] = self.request.session['import_errors']
        return context


class UploadExcel(PermissionRequiredMixin, RevisionMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'crm/manager/client/upload_excel.html'
    permission_required = 'client.add'
    success_url = reverse_lazy('crm:manager:client:import-report')

    def form_valid(self, form):
        file = form.cleaned_data['file']
        ignore_first_row = form.cleaned_data['ignore_first_row']
        name_col = form.cleaned_data['name_col']
        phone_col = form.cleaned_data['phone_col']
        birthday_col = form.cleaned_data['birthday_col']
        vk_col = form.cleaned_data['vk_col']
        balance_col = form.cleaned_data['balance_col']

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
        except Exception:
            form._errors[forms.NON_FIELD_ERRORS] = ErrorList([
                u'Неподдерживаемый формат файла!'
            ])
            return self.form_invalid(form)

        iter_rows = iter(ws.rows)
        if ignore_first_row:
            next(iter_rows)

        skipped = 0
        added = 0
        clients_to_add = []
        vk_domains = []
        errors={}

        for index, row in enumerate(iter_rows):
            try:
                name = row[cell.column_index_from_string(name_col) - 1].value
            except (IndexError, ValueError):
                errors["{}{}".format(name_col, index+1)] = "Ошибка в имени или имя пустое."
                skipped += 1
                continue
            if name == None:
                errors["{}{}".format(name_col, index + 1)] = "Имя пустое."
                skipped += 1
                continue

            try:
                phone = self.try_parse_phone(row[cell.column_index_from_string(phone_col) - 1].value)
            except IndexError:
                phone = Client._meta.get_field('phone_number').get_default()
            except ValueError:
                errors["{}{}".format(phone_col, index+1)] = "Неверный формат номера."
                skipped += 1
                continue

            try:
                birthday = try_parse_date(row[cell.column_index_from_string(birthday_col) - 1].value)
            except (IndexError, TypeError):
                birthday = Client._meta.get_field('birthday').get_default()
            except ValueError:
                errors["{}{}".format(birthday_col, index+1)] = "Неверный формат даты. Допустимые форматы: " + allowed_date_formats_ru;
                skipped += 1
                continue


            try:
                balance = self.try_parse_balance(row[cell.column_index_from_string(balance_col) - 1].value)
            except (IndexError, AttributeError):
                balance = Client._meta.get_field('balance').get_default()
            except (ValueError, TypeError):
                errors["{}{}".format(balance_col, index+1)] = "Неверный формат числа. Допустимые форматы: целые числа, дробные разделенные точкой или запятой."
                skipped += 1
                continue

            try:
                m = re.search(utils.VK_PAGE_REGEXP, row[cell.column_index_from_string(vk_col) - 1].value)
                vk_domain = m.group('user_id')
            except (IndexError, TypeError, AttributeError):
                vk_domain = Client._meta.get_field('vk_user_id').get_default()
            except ValueError:
                errors["{}{}".format(vk_col, index+1)] = "Неверный формат ссылки. Допустимые форматы: vk.com/user_id или https://vk.com/user_id"
                skipped += 1
                continue

            exists = Client.objects.filter(
                name=name,
                phone_number=phone,
                birthday=birthday
            ).exists()

            if not exists:
                client = Client(name = name, phone_number = phone, birthday = birthday)
                if balance:
                    client.balance = balance
                clients_to_add.append(client)
                if vk_domain:
                    vk_domains.append(vk_domain)
                added += 1
            else:
                skipped+=1

            if len(clients_to_add) >= 1000:
                vk_user_ids = get_vk_user_ids(vk_domains)
                for vk_user_id in vk_user_ids:
                    clients_to_add[vk_user_ids.index(vk_user_id)].vk_user_id = vk_user_id
                Client.objects.bulk_create(clients_to_add)
                clients_to_add = []
                vk_domains = []

        vk_user_ids = get_vk_user_ids(vk_domains)
        for i, vk_user_id in enumerate(vk_user_ids):
            clients_to_add[i].vk_user_id = vk_user_id

        Client.objects.bulk_create(clients_to_add)
        self.request.session['import_errors'] = errors

        messages.info(self.request, 'Создано записей: {}. Пропущено записей: {}'.format(added, skipped) )

        return super(UploadExcel, self).form_valid(form)

    def try_parse_phone(self, raw_value):
        phone = re.sub("\D", "", str(raw_value))[0:14]
        return phone

    def try_parse_balance(self, raw_value):
        if isinstance(raw_value, float):
            return  raw_value;
        balance = float(raw_value.replace(',','.'))
        return balance
